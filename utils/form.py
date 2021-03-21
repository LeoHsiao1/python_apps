"""
Use to read and write tables in CSV, XLSX format.
The installation is required for handling XLSX tables: pip install openpyxl
"""


def read_csv(filename, *args, **kwargs):
    """
    从一个csv文件中读取每行数据，返回一个二维列表、
    - 基于csv模块。
    - 该函数的参数列表与open()相同。
    """
    import csv
    result = []
    with open(filename, *args, **kwargs) as f:
        csv_reader = csv.reader(f)   # 在一个文件流f上创建csv阅读器
        try:
            for line in csv_reader:  # 迭代csv_reader的内容
                result.append(line)
        except csv.Error as e:       # 捕捉读取csv文件时的异常
            raise OSError('A reading error in file {}, line {}:\n{}'
                          .format(filename, csv_reader.line_num, e))
    return result


def write_csv(data, filename, mode='w', newline='', **kwargs):
    """ 
    将数据（比如二维列表）转换成csv格式，再保存到指定文件中。
    - 基于csv模块。
    """
    import csv
    with open(filename, mode, newline=newline, **kwargs) as f:
        csv_writer = csv.writer(f)  # 在文件流f上创建一个csv写入器
        csv_writer.writerows(data)  # 写入多行


def read_xlsx(filename, read_only=True, *args, **kwargs):
    """
    读取一个xlsx表格中的全部数据，保存为一个字典返回。
    - 该字典的key为xlsx中一个sheet的名字，value为该sheet的所有行组成的list。
    - 该函数的参数与 openpyxl.load_workbook() 相同。
    - read_only=True 表示以只读模式打开，读取速度更快。
    """
    from openpyxl import load_workbook
    wb = load_workbook(filename, read_only, *args, **kwargs)
    data_dict = {}
    # 遍历xlsx中的每个sheet，遍历每个sheet中的每行数据，保存为字典类型
    for name in wb.sheetnames:
        data_dict[name] = [row for row in wb[name].values]
    wb.close()
    return data_dict


def write_xlsx(data_dict, filename, write_only=True):
    """
    将一个字典写入xlsx表格。
    - 如果输入的data不是字典类型，会先转换成 key="Sheet1" 的字典。
    - 该字典的key为xlsx中一个sheet的名字，value为该sheet的所有行组成的list。
    - write_only=True 表示以只写模式打开，读取速度更快。
    """
    from openpyxl import Workbook

    if not isinstance(data_dict, dict):
        data_dict = {"Sheet1": data_dict}

    wb = Workbook(data_dict, write_only)
    # 遍历data_dict中的每个value，遍历每个value的每行数据，保存为xlsx表格
    for k, v in data_dict.items():
        ws = wb.create_sheet(k)
        for row in v:
            ws.append(row)
    wb.save(filename)
    wb.close()


# sample
if __name__ == '__main__':
    # 读写csv
    data = [[1, 2, 3],
            [4, 5, 6]]
    write_csv(data, '1.csv')
    for line in read_csv('1.csv'):
        print(line)

    # 读写xlsx
    dict_data = {'sheet1': data,
                 'sheet2': data}
    write_xlsx(dict_data, '1.xlsx')
    for k, v in read_xlsx('1.xlsx').items():
        print(k, v)
